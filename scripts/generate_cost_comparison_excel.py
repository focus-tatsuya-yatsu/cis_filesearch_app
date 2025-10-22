#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CIS AWS Cost Comparison Excel Generator
3パターンのコスト比較Excelファイルを生成
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# 定数
EXCHANGE_RATE = 150
OUTPUT_FILE = "../docs/CIS_AWS_Cost_Comparison_3Patterns.xlsx"

# AWS色定義
AWS_ORANGE = "FF9900"
AWS_DARK_BLUE = "232F3E"
PATTERN1_COLOR = "E3F2FD"  # Light Blue
PATTERN2_COLOR = "FFF9C4"  # Light Yellow
PATTERN3_COLOR = "C8E6C9"  # Light Green

def create_header_style():
    """ヘッダー用スタイル"""
    return {
        'font': Font(name='Arial', size=12, bold=True, color="FFFFFF"),
        'fill': PatternFill(start_color=AWS_DARK_BLUE, end_color=AWS_DARK_BLUE, fill_type="solid"),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }

def create_data_style():
    """データセル用スタイル"""
    return {
        'font': Font(name='Arial', size=10),
        'alignment': Alignment(horizontal='left', vertical='center'),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }

def apply_title(ws, title, row=1):
    """タイトル適用"""
    ws.merge_cells(f'A{row}:H{row}')
    cell = ws[f'A{row}']
    cell.value = title
    cell.font = Font(name='Arial', size=16, bold=True, color=AWS_DARK_BLUE)
    cell.alignment = Alignment(horizontal='center', vertical='center')

def apply_subtitle(ws, subtitle, row=2):
    """サブタイトル適用"""
    ws.merge_cells(f'A{row}:H{row}')
    cell = ws[f'A{row}']
    cell.value = subtitle
    cell.font = Font(name='Arial', size=10, italic=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

# ==============================
# Sheet 1: サマリー
# ==============================
def create_summary_sheet(wb):
    """サマリーシート作成"""
    ws = wb.create_sheet("サマリー", 0)

    apply_title(ws, "CIS ファイル検索システム - AWS コスト比較サマリー", 1)
    apply_subtitle(ws, f"作成日: {datetime.now().strftime('%Y年%m月%d日')} | 為替レート: ¥{EXCHANGE_RATE}/USD", 2)

    # データ
    data = [
        ["パターン", "月額 (USD)", "月額 (JPY)", "年額 (USD)", "年額 (JPY)", "3年TCO (USD)", "3年TCO (JPY)", "削減率 vs P2"],
        ["Pattern 1: コスト最優先", 90.40, 13560, 1084.80, 162720, 3254.40, 488160, "-75%"],
        ["Pattern 2: 高可用性+セキュリティ", 1105.50, 165825, 13266.00, 1989900, 39798.00, 5969700, "0%"],
        ["Pattern 3: 月次バッチ (推奨)", 47.24, 7086, 566.88, 85032, 1810.38, 271557, "-96%"],
    ]

    start_row = 4
    for r_idx, row in enumerate(data, start=start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == start_row:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color=AWS_DARK_BLUE, end_color=AWS_DARK_BLUE, fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='right' if isinstance(value, (int, float)) else 'left')
                # パターン別の背景色
                if "Pattern 1" in str(value):
                    cell.fill = PatternFill(start_color=PATTERN1_COLOR, end_color=PATTERN1_COLOR, fill_type="solid")
                elif "Pattern 2" in str(value):
                    cell.fill = PatternFill(start_color=PATTERN2_COLOR, end_color=PATTERN2_COLOR, fill_type="solid")
                elif "Pattern 3" in str(value):
                    cell.fill = PatternFill(start_color=PATTERN3_COLOR, end_color=PATTERN3_COLOR, fill_type="solid")

    # 列幅調整
    ws.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 15

    # 棒グラフ: 月額コスト比較
    chart1 = BarChart()
    chart1.title = "月額コスト比較 (USD)"
    chart1.y_axis.title = "コスト (USD)"
    chart1.x_axis.title = "パターン"

    data_ref = Reference(ws, min_col=2, min_row=4, max_row=7, max_col=2)
    cats_ref = Reference(ws, min_col=1, min_row=5, max_row=7)
    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(cats_ref)
    chart1.width = 15
    chart1.height = 8
    ws.add_chart(chart1, "A10")

    # 棒グラフ: 3年TCO比較
    chart2 = BarChart()
    chart2.title = "3年間TCO比較 (USD)"
    chart2.y_axis.title = "コスト (USD)"
    chart2.x_axis.title = "パターン"

    data_ref2 = Reference(ws, min_col=6, min_row=4, max_row=7, max_col=6)
    chart2.add_data(data_ref2, titles_from_data=True)
    chart2.set_categories(cats_ref)
    chart2.width = 15
    chart2.height = 8
    ws.add_chart(chart2, "J10")

    # 推奨パターン選択フローチャート
    flow_start_row = 26
    ws.merge_cells(f'A{flow_start_row}:H{flow_start_row}')
    ws[f'A{flow_start_row}'].value = "推奨パターン選択フローチャート"
    ws[f'A{flow_start_row}'].font = Font(size=14, bold=True, color=AWS_DARK_BLUE)
    ws[f'A{flow_start_row}'].alignment = Alignment(horizontal='center')

    flow_data = [
        ["条件", "推奨パターン", "理由"],
        ["過去データのみ検索（新データ不要）", "Pattern 3", "最高のコスト効率、VPN97%削減、100万ファイル対応"],
        ["少量ファイル・低頻度検索", "Pattern 1", "基本機能のみで十分、初期投資最小化"],
        ["本格運用・高可用性必須", "Pattern 2", "Multi-AZ、リアルタイム同期、AI機能フル"],
    ]

    flow_start_row += 1
    for r_idx, row in enumerate(flow_data, start=flow_start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == flow_start_row:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color=AWS_DARK_BLUE, end_color=AWS_DARK_BLUE, fill_type="solid")
            cell.alignment = Alignment(horizontal='center' if r_idx == flow_start_row else 'left', vertical='center', wrap_text=True)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 50

# ==============================
# Sheet 2-4: Pattern Details
# ==============================
def create_pattern_detail_sheet(wb, pattern_name, pattern_number, monthly_cost, services_data):
    """パターン詳細シート作成"""
    ws = wb.create_sheet(f"Pattern{pattern_number}_Details")

    apply_title(ws, f"{pattern_name} - 詳細コスト内訳", 1)
    apply_subtitle(ws, f"月額コスト: ${monthly_cost:.2f} (¥{int(monthly_cost * EXCHANGE_RATE):,})", 2)

    # データテーブル
    headers = ["カテゴリ", "サービス", "詳細", "月額 (USD)", "月額 (JPY)", "構成比"]
    start_row = 4

    for c_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=AWS_DARK_BLUE, end_color=AWS_DARK_BLUE, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for r_idx, row in enumerate(services_data, start=start_row+1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal='right' if isinstance(value, (int, float)) else 'left', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # 列幅調整
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12

    # 円グラフ: サービス別構成比
    pie = PieChart()
    pie.title = f"{pattern_name} コスト構成比"

    data_ref = Reference(ws, min_col=4, min_row=start_row, max_row=start_row+len(services_data))
    labels_ref = Reference(ws, min_col=2, min_row=start_row+1, max_row=start_row+len(services_data))
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(labels_ref)
    pie.width = 12
    pie.height = 10
    ws.add_chart(pie, "H5")

    # コストサマリー
    summary_row = start_row + len(services_data) + 3
    ws[f'A{summary_row}'].value = "コストサマリー"
    ws[f'A{summary_row}'].font = Font(size=12, bold=True, color=AWS_DARK_BLUE)

    summary_data = [
        ["月額コスト (USD)", f"${monthly_cost:.2f}"],
        ["月額コスト (JPY)", f"¥{int(monthly_cost * EXCHANGE_RATE):,}"],
        ["年額コスト (USD)", f"${monthly_cost * 12:.2f}"],
        ["年額コスト (JPY)", f"¥{int(monthly_cost * 12 * EXCHANGE_RATE):,}"],
        ["3年TCO (USD)", f"${monthly_cost * 36:.2f}"],
        ["3年TCO (JPY)", f"¥{int(monthly_cost * 36 * EXCHANGE_RATE):,}"],
    ]

    for r_idx, (label, value) in enumerate(summary_data, start=summary_row+1):
        ws[f'A{r_idx}'].value = label
        ws[f'B{r_idx}'].value = value
        ws[f'A{r_idx}'].font = Font(bold=True)
        ws[f'B{r_idx}'].alignment = Alignment(horizontal='right')

# Pattern 1 データ
pattern1_services = [
    ["データベース", "RDS (db.t4g.micro)", "Single-AZ, 20GB gp2", 12.00, 1800, "13.3%"],
    ["検索エンジン", "OpenSearch (t3.small.search)", "1ノード, 10GB", 25.00, 3750, "27.7%"],
    ["コンピューティング", "Lambda", "無料枠内", 0.00, 0, "0.0%"],
    ["データベース", "DynamoDB", "On-Demand, 250MB", 2.00, 300, "2.2%"],
    ["ストレージ", "S3 STANDARD", "50GB", 1.15, 173, "1.3%"],
    ["監視", "CloudWatch Logs", "10GB", 5.00, 750, "5.5%"],
    ["API", "API Gateway", "1.5M リクエスト", 5.25, 788, "5.8%"],
    ["認証", "Cognito", "無料枠内", 0.00, 0, "0.0%"],
    ["ネットワーク", "NAT Gateway", "Single-AZ, 50GB", 35.00, 5250, "38.7%"],
    ["データ転送", "Data Transfer", "50GB", 5.00, 750, "5.5%"],
]

# Pattern 2 データ
pattern2_services = [
    ["データベース", "RDS (db.t4g.medium)", "Multi-AZ, 100GB gp3", 120.00, 18000, "10.9%"],
    ["データベース", "RDS Proxy", "接続プーリング", 15.00, 2250, "1.4%"],
    ["検索エンジン", "OpenSearch (r6g.large×3)", "3-AZ, マスターノード", 450.00, 67500, "40.7%"],
    ["コンピューティング", "Lambda (実行時間)", "4M GB-秒", 85.00, 12750, "7.7%"],
    ["コンピューティング", "Lambda (Provisioned)", "3 PC-時間", 45.00, 6750, "4.1%"],
    ["データベース", "DynamoDB", "Provisioned + On-Demand", 35.00, 5250, "3.2%"],
    ["ストレージ", "S3 (Multi-Tier)", "500GB", 15.00, 2250, "1.4%"],
    ["ストレージ", "S3 CRR", "Cross-Region Replication", 10.00, 1500, "0.9%"],
    ["監視", "CloudWatch", "Logs + Metrics + Alarms", 105.00, 15750, "9.5%"],
    ["API", "API Gateway", "2.5M リクエスト + Cache", 22.00, 3300, "2.0%"],
    ["認証", "Cognito Advanced", "高度なセキュリティ", 2.50, 375, "0.2%"],
    ["CDN", "CloudFront", "200GB", 30.00, 4500, "2.7%"],
    ["キャッシュ", "ElastiCache (Redis)", "t4g.small×2", 35.00, 5250, "3.2%"],
    ["ネットワーク", "NAT Gateway×2", "Multi-AZ, 200GB", 70.00, 10500, "6.3%"],
    ["ロードバランサー", "ALB", "100GB処理", 25.00, 3750, "2.3%"],
    ["セキュリティ", "AWS WAF", "10ルール", 15.00, 2250, "1.4%"],
    ["通知", "Amazon SES", "10,000通", 1.00, 150, "0.1%"],
    ["イベント", "EventBridge", "10ルール", 1.00, 150, "0.1%"],
    ["データ転送", "Data Transfer", "200GB", 20.00, 3000, "1.8%"],
]

# Pattern 3 データ
pattern3_services = [
    ["ネットワーク", "Site-to-Site VPN", "月4時間のみ", 1.20, 180, "2.5%"],
    ["データ転送", "AWS DataSync", "増分20GB/月", 5.00, 750, "10.6%"],
    ["コンピューティング", "Lambda (ARM64)", "63,705 GB-秒", 1.06, 159, "2.2%"],
    ["検索エンジン", "OpenSearch (t3.small)", "50GB gp3, kuromoji+k-NN", 31.57, 4736, "66.8%"],
    ["データベース", "DynamoDB", "5GB storage", 1.26, 189, "2.7%"],
    ["ストレージ", "S3 Intelligent-Tiering", "100GB", 2.15, 323, "4.6%"],
    ["ワークフロー", "Step Functions", "20 transitions", 0.00, 0, "0.0%"],
    ["監視", "CloudWatch", "2GB logs + 10 metrics", 4.00, 600, "8.5%"],
    ["イベント", "EventBridge", "1ルール", 1.00, 150, "2.1%"],
    ["通知", "SNS", "10メッセージ", 0.00, 0, "0.0%"],
]

# ==============================
# Sheet 5: Comparison
# ==============================
def create_comparison_sheet(wb):
    """比較シート作成"""
    ws = wb.create_sheet("Comparison")

    apply_title(ws, "3パターン横並び比較", 1)
    apply_subtitle(ws, "機能・可用性・パフォーマンス・コスト比較", 2)

    # コスト比較
    ws['A4'].value = "コスト比較"
    ws['A4'].font = Font(size=12, bold=True, color=AWS_DARK_BLUE)

    cost_data = [
        ["項目", "Pattern 1", "Pattern 2", "Pattern 3"],
        ["月額 (USD)", "$90.40", "$1,105.50", "$47.24"],
        ["月額 (JPY)", "¥13,560", "¥165,825", "¥7,086"],
        ["年額 (USD)", "$1,084.80", "$13,266.00", "$566.88"],
        ["3年TCO (USD)", "$3,254.40", "$39,798.00", "$1,810.38"],
        ["削減率 vs P2", "-75%", "0%", "-96%"],
    ]

    for r_idx, row in enumerate(cost_data, start=5):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 5:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color=AWS_DARK_BLUE, end_color=AWS_DARK_BLUE, fill_type="solid")
            else:
                if c_idx == 2:
                    cell.fill = PatternFill(start_color=PATTERN1_COLOR, end_color=PATTERN1_COLOR, fill_type="solid")
                elif c_idx == 3:
                    cell.fill = PatternFill(start_color=PATTERN2_COLOR, end_color=PATTERN2_COLOR, fill_type="solid")
                elif c_idx == 4:
                    cell.fill = PatternFill(start_color=PATTERN3_COLOR, end_color=PATTERN3_COLOR, fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # 機能比較
    ws['A13'].value = "機能比較"
    ws['A13'].font = Font(size=12, bold=True, color=AWS_DARK_BLUE)

    feature_data = [
        ["機能", "Pattern 1", "Pattern 2", "Pattern 3"],
        ["基本検索", "✓", "✓", "✓"],
        ["全文検索", "✓ (シンプル)", "✓ (高度)", "✓ (kuromoji)"],
        ["画像類似検索", "✗", "✓", "✓ (k-NN)"],
        ["AI機能", "✗", "✓", "✗"],
        ["リアルタイム同期", "手動", "リアルタイム", "月次バッチ"],
        ["データ保管場所", "AWS", "AWS", "NAS (ハイブリッド)"],
        ["VPN接続", "-", "常時", "月4時間のみ"],
        ["Multi-AZ", "✗", "✓", "✗"],
        ["想定ファイル数", "50,000", "50,000", "1,000,000"],
    ]

    for r_idx, row in enumerate(feature_data, start=14):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 14:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color=AWS_DARK_BLUE, end_color=AWS_DARK_BLUE, fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # 可用性比較
    ws['A25'].value = "可用性比較"
    ws['A25'].font = Font(size=12, bold=True, color=AWS_DARK_BLUE)

    availability_data = [
        ["項目", "Pattern 1", "Pattern 2", "Pattern 3"],
        ["RTO (復旧時間)", "4-8時間", "1-2時間", "4-8時間"],
        ["RPO (復旧時点)", "24時間", "5分", "1ヶ月"],
        ["稼働率 (SLA)", "99%", "99.9%", "99%"],
        ["障害時の影響", "全機能停止", "自動フェイルオーバー", "全機能停止"],
        ["データ同期遅延", "手動", "リアルタイム", "最大1ヶ月"],
    ]

    for r_idx, row in enumerate(availability_data, start=26):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 26:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color=AWS_DARK_BLUE, end_color=AWS_DARK_BLUE, fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # パフォーマンス比較
    ws['A34'].value = "パフォーマンス比較"
    ws['A34'].font = Font(size=12, bold=True, color=AWS_DARK_BLUE)

    performance_data = [
        ["項目", "Pattern 1", "Pattern 2", "Pattern 3"],
        ["平均検索レスポンス", "2-5秒", "0.5-2秒", "0.1-0.2秒"],
        ["ピーク時レスポンス", "5-10秒", "1-3秒", "0.2-0.5秒"],
        ["同時接続数", "10-20ユーザー", "50+ユーザー", "50ユーザー"],
        ["月間検索数想定", "1,000,000回", "1,500,000回", "10,000回"],
    ]

    for r_idx, row in enumerate(performance_data, start=35):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 35:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color=AWS_DARK_BLUE, end_color=AWS_DARK_BLUE, fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # 列幅調整
    ws.column_dimensions['A'].width = 25
    for col in ['B', 'C', 'D']:
        ws.column_dimensions[col].width = 20

# ==============================
# Sheet 6: Scaling
# ==============================
def create_scaling_sheet(wb):
    """スケーリングシート作成"""
    ws = wb.create_sheet("Scaling")

    apply_title(ws, "Pattern 3 スケーリングシナリオ", 1)
    apply_subtitle(ws, "ユーザー数・ファイル数増加時のコスト推移", 2)

    # スケーリングデータ
    scaling_data = [
        ["シナリオ", "ユーザー数", "月間検索数", "ファイル数", "データ量 (GB)", "Lambda (USD)", "OpenSearch (USD)", "月額合計 (USD)", "月額合計 (JPY)", "増加率"],
        ["現在", 50, 10000, 1000000, 500, 1.06, 31.57, 47.24, 7086, "0%"],
        ["小規模拡大", 100, 20000, 1000000, 500, 1.11, 31.57, 48.29, 7244, "2%"],
        ["中規模拡大", 200, 40000, 2000000, 1000, 1.21, 56.88, 73.70, 11055, "56%"],
        ["大規模拡大", 500, 100000, 5000000, 2500, 2.56, 106.00, 123.17, 18476, "161%"],
    ]

    start_row = 4
    for r_idx, row in enumerate(scaling_data, start=start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == start_row:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color=AWS_DARK_BLUE, end_color=AWS_DARK_BLUE, fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='right' if isinstance(value, (int, float)) else 'center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # 列幅調整
    for col, width in [('A', 15), ('B', 12), ('C', 15), ('D', 12), ('E', 15), ('F', 15), ('G', 18), ('H', 18), ('I', 18), ('J', 12)]:
        ws.column_dimensions[col].width = width

    # 折れ線グラフ: ユーザー数増加時のコスト推移
    line_chart = LineChart()
    line_chart.title = "ユーザー数増加時のコスト推移"
    line_chart.x_axis.title = "ユーザー数"
    line_chart.y_axis.title = "月額コスト (USD)"

    data_ref = Reference(ws, min_col=8, min_row=4, max_row=8)
    cats_ref = Reference(ws, min_col=2, min_row=5, max_row=8)
    line_chart.add_data(data_ref, titles_from_data=True)
    line_chart.set_categories(cats_ref)
    line_chart.width = 15
    line_chart.height = 10
    ws.add_chart(line_chart, "A10")

    # スケーリング推奨事項
    rec_row = 28
    ws[f'A{rec_row}'].value = "スケーリング推奨事項"
    ws[f'A{rec_row}'].font = Font(size=12, bold=True, color=AWS_DARK_BLUE)

    recommendations = [
        ["シナリオ", "推奨アクション", "期待効果"],
        ["50→100ユーザー", "現状維持 (t3.small.search)", "コスト増加わずか2%"],
        ["100→200ユーザー", "OpenSearch: t3.medium.search へアップグレード", "検索パフォーマンス2倍向上"],
        ["200→500ユーザー", "OpenSearch: r6g.large.search へアップグレード\n+ Lambda メモリ増強", "大規模検索に対応、レスポンス安定化"],
        ["500ユーザー以上", "Pattern 2 (Multi-AZ) への移行検討", "高可用性確保、ビジネス継続性向上"],
    ]

    for r_idx, row in enumerate(recommendations, start=rec_row+1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == rec_row+1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color=AWS_DARK_BLUE, end_color=AWS_DARK_BLUE, fill_type="solid")
            cell.alignment = Alignment(horizontal='center' if r_idx == rec_row+1 else 'left', vertical='center', wrap_text=True)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 35
    ws.row_dimensions[rec_row+4].height = 30

# ==============================
# Sheet 7: Cost Optimization
# ==============================
def create_optimization_sheet(wb):
    """最適化シート作成"""
    ws = wb.create_sheet("Cost_Optimization")

    apply_title(ws, "コスト最適化提案", 1)
    apply_subtitle(ws, "Reserved Instances、Savings Plans、その他削減施策", 2)

    # Reserved Instances効果
    ws['A4'].value = "1. Reserved Instances / Savings Plans 効果"
    ws['A4'].font = Font(size=12, bold=True, color=AWS_DARK_BLUE)

    ri_data = [
        ["サービス", "現在 (On-Demand)", "1年RI (30%削減)", "3年RI (50%削減)", "年間削減額 (USD)"],
        ["RDS (Pattern 2)", "$1,440", "$1,008", "$720", "$720"],
        ["OpenSearch (Pattern 2)", "$5,400", "$3,780", "$2,700", "$2,700"],
        ["Lambda Compute Savings Plans", "$1,020", "$714", "$510", "$510"],
        ["合計削減額", "-", "$2,502", "$3,930", "$3,930"],
    ]

    for r_idx, row in enumerate(ri_data, start=5):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 5:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color=AWS_DARK_BLUE, end_color=AWS_DARK_BLUE, fill_type="solid")
            elif r_idx == 8:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color=PATTERN3_COLOR, end_color=PATTERN3_COLOR, fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Pattern 3 コスト削減内訳
    ws['A11'].value = "2. Pattern 3 主要コスト削減施策"
    ws['A11'].font = Font(size=12, bold=True, color=AWS_DARK_BLUE)

    reduction_data = [
        ["削減項目", "従来構成", "Pattern 3", "削減額/月 (USD)", "削減率"],
        ["VPN常時接続 → 月4時間", "$36.50", "$1.20", "$34.80", "97%"],
        ["DataSync リアルタイム → 増分", "$20.00", "$5.00", "$15.00", "75%"],
        ["Lambda常時 → 月次バッチ", "$15.00", "$1.06", "$13.94", "93%"],
        ["Lambda x86 → ARM64", "$1.33", "$1.06", "$0.27", "20%"],
        ["S3 STANDARD → Intelligent-Tiering", "$3.35", "$2.15", "$1.20", "36%"],
        ["Multi-AZ → Single-AZ (OpenSearch)", "$481.57", "$31.57", "$450.00", "93%"],
        ["合計削減額", "-", "-", "$515.21", "91%"],
    ]

    for r_idx, row in enumerate(reduction_data, start=12):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 12:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color=AWS_DARK_BLUE, end_color=AWS_DARK_BLUE, fill_type="solid")
            elif r_idx == 18:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color=PATTERN3_COLOR, end_color=PATTERN3_COLOR, fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # その他の最適化施策
    ws['A21'].value = "3. その他のコスト最適化施策"
    ws['A21'].font = Font(size=12, bold=True, color=AWS_DARK_BLUE)

    other_optimizations = [
        ["施策", "説明", "期待削減率", "適用難易度"],
        ["S3 Lifecycle Policy", "90日以上のログをGlacierへ自動移行", "10-15%", "低"],
        ["Lambda SnapStart", "コールドスタート削減でProvisioned Concurrency不要", "15-20%", "中"],
        ["CloudWatch Logs → Athena", "ログ分析をAthenaで実施", "40-50%", "中"],
        ["Spot Instances (開発環境)", "開発環境でSpot利用", "70%", "低"],
        ["Compute Optimizer", "自動リソース最適化提案", "5-10%", "低"],
        ["Trusted Advisor", "コスト最適化チェック", "変動", "低"],
    ]

    for r_idx, row in enumerate(other_optimizations, start=22):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 22:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color=AWS_DARK_BLUE, end_color=AWS_DARK_BLUE, fill_type="solid")
            cell.alignment = Alignment(horizontal='center' if r_idx == 22 else 'left', vertical='center', wrap_text=True)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # 列幅調整
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15

    # 最適化実施ロードマップ
    ws['A31'].value = "4. 最適化実施ロードマップ"
    ws['A31'].font = Font(size=12, bold=True, color=AWS_DARK_BLUE)

    roadmap = [
        ["フェーズ", "期間", "施策", "期待削減率"],
        ["Phase 1: 即時実施", "Week 1-2", "S3 Lifecycle Policy設定\nCloudWatch Logs保存期間最適化", "5%"],
        ["Phase 2: 短期", "Month 1-3", "Reserved Instances購入 (1年)\nLambda ARM64移行", "15%"],
        ["Phase 3: 中期", "Month 4-6", "Savings Plans検討\nCompute Optimizer導入", "10%"],
        ["Phase 4: 長期", "Month 7-12", "3年RI検討\nアーキテクチャ最適化", "15%"],
        ["合計削減効果", "-", "-", "40-45%"],
    ]

    for r_idx, row in enumerate(roadmap, start=32):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 32:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color=AWS_DARK_BLUE, end_color=AWS_DARK_BLUE, fill_type="solid")
            elif r_idx == 36:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color=PATTERN3_COLOR, end_color=PATTERN3_COLOR, fill_type="solid")
            cell.alignment = Alignment(horizontal='center' if r_idx == 32 else 'left', vertical='center', wrap_text=True)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.row_dimensions[33].height = 30
    ws.row_dimensions[34].height = 30
    ws.row_dimensions[35].height = 30
    ws.row_dimensions[36].height = 30

# ==============================
# メイン実行
# ==============================
def main():
    print("CIS AWS Cost Comparison Excel Generator")
    print("=" * 60)

    # Workbook作成
    wb = Workbook()
    wb.remove(wb.active)  # デフォルトシート削除

    print("📊 Sheet 1: サマリー作成中...")
    create_summary_sheet(wb)

    print("📋 Sheet 2: Pattern 1 詳細作成中...")
    create_pattern_detail_sheet(wb, "Pattern 1: コスト最優先", 1, 90.40, pattern1_services)

    print("📋 Sheet 3: Pattern 2 詳細作成中...")
    create_pattern_detail_sheet(wb, "Pattern 2: 高可用性+セキュリティ", 2, 1105.50, pattern2_services)

    print("📋 Sheet 4: Pattern 3 詳細作成中...")
    create_pattern_detail_sheet(wb, "Pattern 3: 月次バッチ (推奨)", 3, 47.24, pattern3_services)

    print("🔄 Sheet 5: 比較表作成中...")
    create_comparison_sheet(wb)

    print("📈 Sheet 6: スケーリングシナリオ作成中...")
    create_scaling_sheet(wb)

    print("💡 Sheet 7: コスト最適化提案作成中...")
    create_optimization_sheet(wb)

    # 保存
    print(f"\n💾 保存中: {OUTPUT_FILE}")
    wb.save(OUTPUT_FILE)

    print("✅ Excel生成完了!")
    print(f"📁 出力ファイル: {OUTPUT_FILE}")
    print("=" * 60)

if __name__ == "__main__":
    main()
